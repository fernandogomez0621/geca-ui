"""GPU Worker for YOLO training, inference, and video annotation"""
from fastapi import FastAPI
from pydantic import BaseModel
import os, json, time, subprocess, threading
import numpy as np
import cv2
import torch

app = FastAPI()
SHARED = os.getenv("SHARED_DIR", "/mnt/shared")
tasks = {}  # task_id -> progress dict


class TrainRequest(BaseModel):
    dataset_name: str
    model_base: str = "yolo26m.pt"
    experiment_name: str = ""
    epochs: int = 100
    batch: int = 4
    patience: int = 50
    imgsz: int = 640
    freeze: int = 0
    lr0: float = 0.01
    cos_lr: bool = False
    mixup: float = 0.0
    copy_paste: float = 0.0
    scale: float = 0.5
    cls: float = 0.5


class InferenceRequest(BaseModel):
    model_name: str
    video_path: str
    fps_process: int = 10
    conf: float = 0.25
    batch_size: int = 32


class VideoRequest(BaseModel):
    model_name: str
    video_path: str
    resolution: int = 480
    conf: float = 0.25
    batch_size: int = 32
    crf: int = 28


@app.get("/health")
def health():
    gpu = torch.cuda.get_device_name(0) if torch.cuda.is_available() else "N/A"
    return {"status": "ok", "gpu": gpu, "cuda": torch.cuda.is_available()}


@app.get("/tasks/{task_id}")
def get_task(task_id: str):
    return tasks.get(task_id, {"status": "not_found"})


# ============ TRAINING ============

@app.post("/train")
def start_training(req: TrainRequest):
    task_id = f"train_{int(time.time())}"
    tasks[task_id] = {"status": "starting", "type": "train", "progress": 0, "epoch": 0, "total_epochs": req.epochs}
    t = threading.Thread(target=run_training, args=(task_id, req), daemon=True)
    t.start()
    return {"task_id": task_id, "status": "started"}


def run_training(task_id, req):
    try:
        from ultralytics import YOLO

        model_path = os.path.join(SHARED, "models", req.model_base)
        yaml_path = os.path.join(SHARED, "datasets", "ready", req.dataset_name, "data.yaml")
        runs_dir = os.path.join(SHARED, "runs", "detect")
        exp_name = req.experiment_name or f"{req.dataset_name}_v1"

        if not os.path.exists(model_path):
            tasks[task_id] = {"status": "error", "message": f"Modelo {req.model_base} no encontrado"}
            return
        if not os.path.exists(yaml_path):
            tasks[task_id] = {"status": "error", "message": f"Dataset {req.dataset_name} no encontrado"}
            return

        tasks[task_id].update({"status": "loading_model", "experiment": exp_name})

        model = YOLO(model_path)

        # Custom callback to track progress
        def on_train_epoch_end(trainer):
            epoch = trainer.epoch + 1
            metrics = trainer.metrics or {}
            tasks[task_id].update({
                "status": "training",
                "epoch": epoch,
                "total_epochs": req.epochs,
                "progress": round(epoch / req.epochs * 100, 1),
                "box_loss": round(float(trainer.loss_items[0]), 4) if trainer.loss_items is not None else None,
                "cls_loss": round(float(trainer.loss_items[1]), 4) if trainer.loss_items is not None else None,
                "mAP50": round(float(metrics.get("metrics/mAP50(B)", 0)), 4),
                "mAP50_95": round(float(metrics.get("metrics/mAP50-95(B)", 0)), 4),
            })

        model.add_callback("on_train_epoch_end", on_train_epoch_end)

        tasks[task_id].update({"status": "training", "epoch": 0})

        results = model.train(
            data=yaml_path, epochs=req.epochs, imgsz=req.imgsz, batch=req.batch,
            device=0, patience=req.patience, name=exp_name, project=runs_dir,
            exist_ok=True, plots=True, mosaic=1.0, mixup=req.mixup,
            copy_paste=req.copy_paste, scale=req.scale, fliplr=0.5,
            hsv_h=0.015, hsv_s=0.7, hsv_v=0.4,
            freeze=req.freeze if req.freeze > 0 else None,
            lr0=req.lr0, cos_lr=req.cos_lr, cls=req.cls,
        )

        # Save best model
        best_path = os.path.join(runs_dir, exp_name, "weights", "best.pt")
        output_name = f"geca_{exp_name}_best.pt"
        output_path = os.path.join(SHARED, "models", output_name)
        if os.path.exists(best_path):
            import shutil
            shutil.copy(best_path, output_path)

        # Get final metrics
        best_model = YOLO(best_path)
        metrics = best_model.val()

        tasks[task_id] = {
            "status": "done", "type": "train", "progress": 100,
            "epoch": req.epochs, "total_epochs": req.epochs,
            "experiment": exp_name,
            "model_saved": output_name,
            "mAP50": round(float(metrics.box.map50), 4),
            "mAP50_95": round(float(metrics.box.map), 4),
        }

    except Exception as e:
        tasks[task_id] = {"status": "error", "type": "train", "message": str(e)}


# ============ INFERENCE ============

@app.post("/inference")
def start_inference(req: InferenceRequest):
    task_id = f"inference_{int(time.time())}"
    tasks[task_id] = {"status": "starting", "type": "inference", "current": 0, "total": 0}
    t = threading.Thread(target=run_inference, args=(task_id, req), daemon=True)
    t.start()
    return {"task_id": task_id, "status": "started"}


def run_inference(task_id, req):
    try:
        from ultralytics import YOLO
        from collections import defaultdict
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        model_path = os.path.join(SHARED, "models", req.model_name)
        model = YOLO(model_path)
        class_names = model.names

        cap = cv2.VideoCapture(req.video_path)
        video_fps = cap.get(cv2.CAP_PROP_FPS)
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        frame_interval = max(1, int(video_fps / req.fps_process))
        total_to_process = total_frames // frame_interval
        video_name = os.path.splitext(os.path.basename(req.video_path))[0]

        tasks[task_id].update({"status": "processing", "total": total_to_process, "video": video_name})

        class_det = defaultdict(int)
        class_fc = defaultdict(int)
        processed = 0
        idx = 0
        batch = []

        def run_batch(b):
            nonlocal processed
            if not b: return
            for r in model(b, conf=req.conf, imgsz=640, verbose=False):
                fc = set()
                for box in r.boxes:
                    cn = class_names.get(int(box.cls.item()), "?")
                    class_det[cn] += 1
                    fc.add(cn)
                for cn in fc:
                    class_fc[cn] += 1
            processed += len(b)
            tasks[task_id].update({"current": processed, "annotations": sum(class_det.values())})

        while True:
            ret, frame = cap.read()
            if not ret: break
            if idx % frame_interval == 0:
                batch.append(frame)
                if len(batch) >= req.batch_size:
                    run_batch(batch)
                    batch = []
            idx += 1
        run_batch(batch)
        cap.release()

        # Generate Excel
        output_dir = os.path.join(SHARED, "results")
        os.makedirs(output_dir, exist_ok=True)

        metrics_list = []
        for cn in sorted(class_det, key=lambda x: -class_det[x]):
            td, fw = class_det[cn], class_fc[cn]
            metrics_list.append({
                "Etiqueta": cn, "Total detecciones": td, "Frames con deteccion": fw,
                "Media cuando aparece": round(td / fw, 6) if fw else 0,
                "Media total": round(td / processed, 6) if processed else 0,
                "Tiempo pantalla (s)": round(fw / req.fps_process, 1),
                "Porcentaje tiempo (%)": round(fw / processed * 100, 6) if processed else 0,
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Métricas"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="2E3440", end_color="2E3440", fill_type="solid")
        b = Border(left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
                   top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"))
        headers = ["Etiqueta", "Total detecciones", "Frames con deteccion", "Media cuando aparece",
                   "Media total", "Tiempo pantalla (s)", "Porcentaje tiempo (%)"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = hf, hfill, Alignment(horizontal="center"), b
        for ri, m in enumerate(metrics_list, 2):
            for col, key in enumerate(headers, 1):
                c = ws.cell(row=ri, column=col, value=m[key])
                c.border = b
        ws.column_dimensions["A"].width = 20
        for col in "BCDEFG":
            ws.column_dimensions[col].width = 22
        excel_path = os.path.join(output_dir, f"{video_name}_metrics.xlsx")
        wb.save(excel_path)

        # Generate presencia chart
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        labels = [m["Etiqueta"] for m in metrics_list]
        pcts = [m["Porcentaje tiempo (%)"] for m in metrics_list]
        colors = ["#6c5ce7", "#00b894", "#e17055", "#ffd43b", "#3b82f6", "#ef4444", "#10b981", "#f59e0b"]
        fig, ax = plt.subplots(figsize=(10, max(3, len(labels) * 0.6)))
        bars = ax.barh(labels, pcts, color=[colors[i % len(colors)] for i in range(len(labels))])
        ax.set_xlabel("Tiempo en pantalla (%)")
        ax.set_title(f"Presencia de Marcas — {video_name}", fontweight="bold")
        ax.invert_yaxis()
        for bar, p in zip(bars, pcts):
            ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height() / 2, f"{p:.1f}%", va="center", fontsize=9)
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, f"{video_name}_presencia.png"), dpi=150, bbox_inches="tight")
        plt.close()

        tasks[task_id] = {
            "status": "done", "type": "inference", "current": processed, "total": total_to_process,
            "video": video_name, "annotations": sum(class_det.values()),
            "excel": f"{video_name}_metrics.xlsx", "metrics": metrics_list,
        }

    except Exception as e:
        tasks[task_id] = {"status": "error", "type": "inference", "message": str(e)}


# ============ VIDEO ANOTADO ============

@app.post("/video-annotate")
def start_video_annotate(req: VideoRequest):
    task_id = f"video_{int(time.time())}"
    tasks[task_id] = {"status": "starting", "type": "video", "current": 0, "total": 0}
    t = threading.Thread(target=run_video_annotate, args=(task_id, req), daemon=True)
    t.start()
    return {"task_id": task_id, "status": "started"}


def run_video_annotate(task_id, req):
    try:
        from ultralytics import YOLO
        from collections import defaultdict

        model_path = os.path.join(SHARED, "models", req.model_name)
        model = YOLO(model_path)
        class_names = model.names

        cap = cv2.VideoCapture(req.video_path)
        orig_fps = cap.get(cv2.CAP_PROP_FPS)
        orig_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        orig_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        video_name = os.path.splitext(os.path.basename(req.video_path))[0]

        ratio = min(1, req.resolution / orig_h)
        out_w = int(orig_w * ratio)
        out_w = out_w if out_w % 2 == 0 else out_w + 1
        out_h = int(orig_h * ratio)
        out_h = out_h if out_h % 2 == 0 else out_h + 1

        output_dir = os.path.join(SHARED, "results")
        os.makedirs(output_dir, exist_ok=True)
        temp_path = os.path.join(output_dir, f"{video_name}_temp.mp4")
        output_path = os.path.join(output_dir, f"{video_name}_anotado_{req.resolution}p.mp4")

        COLORS = [(108, 92, 231), (0, 184, 148), (225, 112, 85), (255, 212, 59),
                  (59, 130, 246), (239, 68, 68), (16, 185, 129), (245, 158, 11)]

        writer = cv2.VideoWriter(temp_path, cv2.VideoWriter_fourcc(*"mp4v"), orig_fps, (out_w, out_h))

        tasks[task_id].update({"status": "processing", "total": total_frames, "video": video_name})

        class_det = defaultdict(int)
        class_fc = defaultdict(int)
        processed = 0
        batch_frames = []

        def draw_and_write(batch_list):
            nonlocal processed
            if not batch_list: return
            results = model(batch_list, conf=req.conf, imgsz=640, verbose=False)
            sx, sy = out_w / orig_w, out_h / orig_h
            for result, frame in zip(results, batch_list):
                out = cv2.resize(frame, (out_w, out_h), interpolation=cv2.INTER_AREA) if ratio < 1 else frame.copy()
                fc = set()
                for box in result.boxes:
                    cid = int(box.cls.item())
                    conf_val = float(box.conf.item())
                    x1, y1, x2, y2 = [int(v) for v in (box.xyxy[0].cpu().numpy() * [sx, sy, sx, sy])]
                    color = COLORS[cid % len(COLORS)]
                    cn = class_names.get(cid, str(cid))
                    cv2.rectangle(out, (x1, y1), (x2, y2), color, 2)
                    label = f"{cn} {conf_val:.0%}"
                    (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 1)
                    cv2.rectangle(out, (x1, y1 - th - 8), (x1 + tw + 4, y1), color, -1)
                    cv2.putText(out, label, (x1 + 2, y1 - 4), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
                    class_det[cn] += 1
                    fc.add(cn)
                for cn in fc:
                    class_fc[cn] += 1
                writer.write(out)
            processed += len(batch_list)
            tasks[task_id].update({"current": processed, "progress": round(processed / total_frames * 100, 1)})

        while True:
            ret, frame = cap.read()
            if not ret: break
            batch_frames.append(frame)
            if len(batch_frames) >= req.batch_size:
                draw_and_write(batch_frames)
                batch_frames = []
        draw_and_write(batch_frames)
        cap.release()
        writer.release()

        # Compress to H.264
        tasks[task_id].update({"status": "compressing"})
        subprocess.run(["ffmpeg", "-i", temp_path, "-c:v", "libx264", "-crf", str(req.crf),
                        "-preset", "fast", "-movflags", "+faststart", "-y", output_path],
                       capture_output=True)
        if os.path.exists(temp_path):
            os.remove(temp_path)

        mb = os.path.getsize(output_path) / 1024 / 1024

        tasks[task_id] = {
            "status": "done", "type": "video", "current": total_frames, "total": total_frames,
            "progress": 100, "video": video_name,
            "output": f"{video_name}_anotado_{req.resolution}p.mp4",
            "size_mb": round(mb, 1),
        }

    except Exception as e:
        tasks[task_id] = {"status": "error", "type": "video", "message": str(e)}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8002)
